import pymysql.cursors
from openpyxl import load_workbook
from slugify import slugify

connection = pymysql.connect(host='localhost',
                             user='root',
                             password='11/*aa',
                             db='pttpk',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
cursor = connection.cursor()

try:
    wb = load_workbook('data/pk_list_29.04.2016.xlsx')
    ws = wb.active

    regions = {}
    for index, row in enumerate(ws.rows):
        if index == 0:
            continue

        index_str = str(index + 1)

        city = ws["A" + index_str].value
        district = ws["B" + index_str].value
        neighborhood = ws["C" + index_str].value
        part = ws["D" + index_str].value
        postal_code = ws["E" + index_str].value

        if city not in regions:
            regions[city] = {}

        if district not in regions[city]:
            regions[city][district] = {}

        if neighborhood not in regions[city][district]:
            regions[city][district][neighborhood] = {}

        regions[city][district][neighborhood][part] = postal_code

    for city in regions.iterkeys():
        city_name = city.title()
        city_slug = slugify(city_name)

        with connection.cursor() as cursor:
            sql = "INSERT INTO `city` (`name`, `slug`) VALUES (%s, %s) RETURNING id;"
            cursor.execute(sql, (city_name, city_slug))
            city_id = cursor.insert_id()

            for district in regions[city].iterkeys():
                district_name = district.title()
                district_slug = slugify(district_name)

                with connection.cursor() as cursor:
                    sql = "INSERT INTO `district` (`city_id`, `name`, `slug`) VALUES (%s, %s, %s) RETURNING id;"
                    cursor.execute(sql, (city_id, district_name, district_slug))
                    district_id = cursor.insert_id()

                    for neighborhood in regions[city][district].iterkeys():
                        neighborhood_name = neighborhood.title()
                        neighborhood_slug = slugify(neighborhood_name)

                        with connection.cursor() as cursor:
                            sql = "INSERT INTO `neighborhood` (`district_id`, `name`, `slug`) VALUES (%s, %s, %s) RETURNING id;"
                            cursor.execute(sql, (district_id, neighborhood_name, neighborhood_slug))
                            neighborhood_id = cursor.insert_id()

                            for part in regions[city][district][neighborhood].iterkeys():
                                part_name = part.title()
                                part_slug = slugify(part_name)

                                with connection.cursor() as cursor:
                                    sql = "INSERT INTO `part` (`neighborhood_id`, `name`, `slug`, `postal_code`) VALUES (%s, %s, %s, %s) RETURNING id;"
                                    cursor.execute(sql, (neighborhood_id, part_name, part_slug, regions[city][district][neighborhood][part]))
                                    part_id = cursor.insert_id()

        connection.commit()

    '''
    with connection.cursor() as cursor:
        sql = "INSERT INTO `users` (`email`, `password`) VALUES (%s, %s) RETURNING id;"
        cursor.execute(sql, ('webmaster@python.org', 'very-secret'))

    connection.commit()
    '''
finally:
    connection.close()

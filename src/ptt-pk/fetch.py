import pymysql.cursors
from openpyxl import load_workbook
from slugify import slugify

connection = pymysql.connect(host='localhost',
                             user='root',
                             password='11/*aa',
                             db='pttdata',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
cursor = connection.cursor()

try:
    wb = load_workbook('data/pk_20220810.xlsx')
    ws = wb.active

    regions = {}
    for index, row in enumerate(ws.rows):
        if index == 0:
            continue

        index_str = str(index + 1)

        city = ws["A" + index_str].value
        district = ws["B" + index_str].value
        neighborhood = ws["C" + index_str].value
        part = ws["D" + index_str].value
        postal_code = ws["E" + index_str].value

        if city not in regions:
            regions[city] = {}

        if district not in regions[city]:
            regions[city][district] = {}

        if neighborhood not in regions[city][district]:
            regions[city][district][neighborhood] = {}

        regions[city][district][neighborhood][part] = postal_code

    for city, v1 in regions.items():
        print(city)
        city_name = city.strip()
        city_slug = slugify(city_name)

        with connection.cursor() as cursor:
            sql = "INSERT INTO `city` (`name`, `slug`) VALUES (%s, %s);"
            cursor.execute(sql, (city_name, city_slug))
            connection.commit()
            city_id = cursor.lastrowid

            for district, v2 in regions[city].items():
                district_name = district.strip()
                district_slug = slugify(district_name)

                with connection.cursor() as cursor:
                    sql = "INSERT INTO `district` (`city_id`, `name`, `slug`) VALUES (%s, %s, %s);"
                    cursor.execute(sql, (city_id, district_name, district_slug))
                    connection.commit()
                    district_id = cursor.lastrowid

                    for neighborhood, v3 in regions[city][district].items():
                        neighborhood_name = neighborhood.strip()
                        neighborhood_slug = slugify(neighborhood_name)

                        with connection.cursor() as cursor:
                            sql = "INSERT INTO `neighborhood` (`district_id`, `name`, `slug`) VALUES (%s, %s, %s);"
                            cursor.execute(sql, (district_id, neighborhood_name, neighborhood_slug))
                            connection.commit()
                            neighborhood_id = cursor.lastrowid

                            for part, v4 in regions[city][district][neighborhood].items():
                                part_name = part.strip()
                                part_slug = slugify(part_name)

                                with connection.cursor() as cursor:
                                    sql = "INSERT INTO `part` (`neighborhood_id`, `name`, `slug`, `postal_code`) VALUES (%s, %s, %s, %s);"
                                    cursor.execute(sql, (neighborhood_id, part_name, part_slug, regions[city][district][neighborhood][part]))
                                    connection.commit()
                                    part_id = cursor.lastrowid
                                    postal_code = regions[city][district][neighborhood][part]

                                    print(
                                        city_name + ' / ' + district_name + ' / ' + neighborhood_name + ' / ' + part_name + ' - ' + str(
                                            postal_code))

finally:
    connection.close()
